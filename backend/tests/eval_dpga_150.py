"""
eval_dpga_150.py — live, cacheable end-to-end evaluation of the UNSDG classifier.

Answers, against real ground truth from dpgs.csv.xlsx (141 labeled DPGs) plus a
150-project aux sample from OpenSustain.tech:

  1. Per-SDG precision / recall / F1 and overall micro/macro averages
  2. Reliability of the scores (ROC-AUC per SDG, calibration, label imbalance)
  3. Whether the developer's thresholds are good:
       - main() uses            threshold = 0.55, top_k = 10, alpha = 0.7
       - classify_repo()        threshold = 0.55
       - app.py /api/classify_st_url filters predictions > 0.4
       - COSINE_LOW / COSINE_HIGH = 0.27 / 0.34 in embedding_similarity_scores
       A threshold + alpha sweep on the labeled set finds the optimal operating
       point and reports how far the developer's defaults sit from it.
  4. Whether the classification pipeline actually uses the LLM summary (the
     harness re-derives the scores from the summary the way classify_repo does
     and verifies they match bit-for-bit), plus an ablation against raw README.
  5. Whether the LLM summaries are relevant and not hallucinated (embedding
     relevance vs README, content-word overlap, numeric-claim agreement,
     structural conformance to the summariser prompt, banned-leakage scan).

Dependencies (all already in the repo's venv): numpy, requests, scikit-learn,
openpyxl, python-dotenv, plus the backend/model deps of embedding_url.py.

Usage
-----
    # 1. make sure the GE-Lab microservice is up on :9010 (or let us boot it):
    python eval_dpga_150.py --boot-models --dpgs 141 --opensustain 150
    # 2. quick smoke run (2 labeled + 5 aux) to validate the wiring:
    python eval_dpga_150.py --boot-models --dpgs 2 --opensustain 5 --smoke
    # 3. re-run from disk cache only (no network / no model), fast:
    python eval_dpga_150.py --offline --dpgs 141 --opensustain 150

Every fetched README, Groq summary, zero-shot and embedding score is cached in
--cache-dir (default backend/.eval_cache) keyed by content hashes, so the first
run is the only slow one. After each project that actually hit a live service
(network fetch, Groq, or the microservice) the harness sleeps --delay seconds
(default 10) to respect forge API / Groq rate limits; fully-cached projects skip
the sleep, so --offline re-runs stay fast. Results land in the cache dir as
report_<timestamp>.json plus predictions_<timestamp>.csv.
"""

from __future__ import annotations

import argparse
import atexit
import hashlib
import json
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from urllib.parse import urlparse

import numpy as np
import requests
from dotenv import load_dotenv

BACKEND_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = BACKEND_DIR.parent
CACHE_DEFAULT = BACKEND_DIR / ".eval_cache"

for p in (str(BACKEND_DIR), str(BACKEND_DIR / "services")):
    if p not in sys.path:
        sys.path.insert(0, p)

load_dotenv(BACKEND_DIR / ".env")

import sdg_constants
import embedding_url as eu
from services.repo_fetcher import (
    ProviderError,
    get_provider as _real_get_provider,
    _DOMAIN_MAP,
)

N_SDGS = 17
DEV_THRESHOLD_MAIN = 0.55
DEV_THRESHOLD_CLASSIFY = 0.55
DEV_THRESHOLD_APP = 0.4
DEV_ALPHA = 0.7
DEV_COSINE_LOW = 0.27
DEV_COSINE_HIGH = 0.34
DEV_TOP_K = 10

_DPG_COL_OFFSET = 3
_SDG_NAMES = sdg_constants.SDG_NAMES
_SUPPORTED_HOSTS = {h.removeprefix("www.") for h in _DOMAIN_MAP}

_STOPWORDS = {
    "a", "an", "the", "and", "or", "but", "for", "of", "to", "in", "on", "with",
    "that", "this", "these", "those", "from", "into", "across", "within", "which",
    "who", "whose", "its", "it", "as", "at", "by", "is", "are", "was", "were",
    "be", "been", "being", "has", "have", "had", "do", "does", "did", "will",
    "would", "can", "could", "should", "may", "might", "must", "not", "no", "so",
    "also", "over", "under", "through", "while", "when", "where", "their", "them",
    "they", "we", "you", "your", "our", "us", "i", "he", "she", "all", "any",
    "both", "each", "few", "more", "most", "other", "some", "such", "than", "too",
    "very", "about", "after", "before", "during", "between", "among", "how", "why",
    "what", "make", "making", "made", "use", "using", "used", "helps", "help",
}

_TECH_NOISE = [
    "python", "javascript", "typescript", "react", "vue", "angular", "node",
    "docker", "kubernetes", "postgres", "postgresql", "mysql", "mongodb",
    "redis", "django", "flask", "fastapi", "java", "kotlin", "swift", "golang",
    "rust", "c++", "ruby", "php", "api", "sdk", "cli", "database", "server",
    "framework", "library", "repository", "readme", "github", "gitlab", "docker",
    "container", "kubernetes", "apache", "nginx", "linux", "tensorflow", "pytorch",
    "javascript", "css", "html", "json", "yaml", "browser", "app",
]

_NUMBER_RE = re.compile(r"\d[\d,.]*")
_SDG_LEAK_RE = re.compile(r"\bSDG\s*\d|\bSustainable Development Goal|\bGoal\s+\d+\b", re.IGNORECASE)
_NO_SIGNAL = "NO_SDG_SIGNAL"


def log(msg: str) -> None:
    print(msg, flush=True)


def now() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


# ─────────────────────────────── data loading ────────────────────────────────

def load_dpgs(path: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    projects = []
    for r in rows:
        if not r or not r[0]:
            continue
        gt = [int(r[_DPG_COL_OFFSET + i] or 0) for i in range(N_SDGS)]
        projects.append({
            "name": str(r[0]).strip(),
            "url": str(r[1] or "").strip(),
            "desc": str(r[2] or "").strip(),
            "gt": gt,
        })
    projects = [p for p in projects if p["url"]]
    return projects


def _norm_key(url: str) -> str:
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    parts = [x for x in parsed.path.strip("/").split("/") if x]
    owner = parts[0] if parts else ""
    repo = parts[1].removesuffix(".git") if len(parts) > 1 else ""
    return f"{host}/{owner}/{repo}"


def load_opensustain(path: Path, n: int, seed: int, exclude_keys: set[str]) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    seen: set[str] = set()
    candidates = []
    for r in rows:
        url = str(r[1] or "").strip()
        name = str(r[0] or "").strip()
        desc = str(r[3] or "").strip()
        if not url:
            continue
        host = (urlparse(url).hostname or "").lower().removeprefix("www.")
        if host not in _SUPPORTED_HOSTS:
            continue
        key = _norm_key(url)
        if key in exclude_keys or key in seen:
            continue
        seen.add(key)
        candidates.append({"name": name, "url": url, "desc": desc, "gt": None})
    rng = random.Random(seed)
    rng.shuffle(candidates)
    return candidates[:n]


# ─────────────────────────────── caching layer ───────────────────────────────

class DiskCache:
    def __init__(self, root: Path):
        self.root = Path(root)
        self.raw = self.root / "raw"
        self.summary = self.root / "summary"
        self.zs = self.root / "zero_shot"
        self.es = self.root / "embedding"
        for d in (self.raw, self.summary, self.zs, self.es):
            d.mkdir(parents=True, exist_ok=True)

    def _path(self, sub: Path, key: str) -> Path:
        return sub / (key + ".json")

    def get(self, sub: Path, key: str):
        p = self._path(sub, key)
        if p.exists():
            return json.loads(p.read_text(encoding="utf-8"))
        return None

    def put(self, sub: Path, key: str, value) -> None:
        self._path(sub, key).write_text(
            json.dumps(value, ensure_ascii=False), encoding="utf-8"
        )

    @staticmethod
    def key(*parts) -> str:
        raw = "|".join(str(p) for p in parts)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]


_cache: DiskCache | None = None
_summary_by_url: dict[str, str] = {}
_live_call_made = False
_orig_zero_shot = eu.zero_shot_scores
_orig_embedding = eu.embedding_similarity_scores
_orig_fetch_text = eu.fetch_repo_text


def _get_token(url: str) -> str | None:
    host = urlparse(url).hostname or ""
    return os.environ.get("GITHUB_TOKEN") if "github.com" in host else None


def _get_raw_sources(url: str, max_retries: int = 1) -> dict:
    ck = _cache.key(url)
    cached = _cache.get(_cache.raw, ck)
    if cached is not None:
        return cached

    global _live_call_made
    _live_call_made = True
    provider = _real_get_provider(url, token=_get_token(url))
    meta: dict = {"name": "", "description": "", "homepage": ""}
    topics: list = []
    readme: str = ""

    def safe(fn, default):
        for attempt in range(max_retries + 1):
            try:
                return fn()
            except ProviderError as e:
                if attempt < max_retries and "rate" in str(e).lower():
                    time.sleep(30)
                    continue
                return default
        return default

    m = safe(provider.fetch_meta, {})
    if isinstance(m, dict):
        meta = m
    topics = safe(provider.fetch_topics, []) or []
    readme = safe(provider.fetch_readme, "") or ""

    raw = {
        "owner": provider._owner,
        "repo": provider._repo,
        "name": meta.get("name") or "",
        "repo_description": meta.get("description") or "",
        "homepage": meta.get("homepage") or "",
        "topics": list(topics),
        "readme": readme,
    }
    _cache.put(_cache.raw, ck, raw)
    return raw


def _cached_summarize(readme: str, name: str, description: str, topics: list[str]) -> str:
    ck = _cache.key(name, description, sorted(topics), readme[:8000])
    cached = _cache.get(_cache.summary, ck)
    if cached is not None:
        return cached.get("summary", "")
    global _live_call_made
    _live_call_made = True
    summary = eu.summarize_for_sdg(
        readme=readme, name=name, description=description, topics=topics
    )
    _cache.put(_cache.summary, ck, {"summary": summary, "ts": now()})
    return summary


def _cached_fetch_repo_text(url: str, project_description: str = "", max_issues: int = 10) -> dict:
    raw = _get_raw_sources(url)
    name = raw["name"]
    description = project_description.strip() or raw["repo_description"] or ""
    topics = raw["topics"]
    readme = raw["readme"]
    text = _cached_summarize(
        readme=readme, name=name, description=description, topics=topics
    )
    _summary_by_url[url] = text
    return {
        "owner": raw["owner"],
        "repo": raw["repo"],
        "text": text,
        "meta": {
            "name": name,
            "description": description,
            "topics": topics,
            "homepage": raw["homepage"],
        },
    }


def _cached_zero_shot(text: str, labels: list[str]):
    ck = _cache.key("zs", text[:6000])
    cached = _cache.get(_cache.zs, ck)
    if cached is not None:
        return np.array(cached["scores"], dtype=float), {
            "labels": labels,
            "scores": cached["scores"],
            "sequence": text[:500],
        }
    global _live_call_made
    _live_call_made = True
    zs, details = _orig_zero_shot(text, labels)
    _cache.put(_cache.zs, ck, {"scores": list(np.asarray(zs, dtype=float))})
    return np.asarray(zs, dtype=float), {
        "labels": labels,
        "scores": list(np.asarray(zs, dtype=float)),
        "sequence": text[:500],
    }


def _cached_embedding(text: str, label_texts: list[str]) -> np.ndarray:
    ck = _cache.key("es", text[:6000])
    cached = _cache.get(_cache.es, ck)
    if cached is not None:
        return np.array(cached["scores"], dtype=float)
    global _live_call_made
    _live_call_made = True
    es = _orig_embedding(text, label_texts)
    _cache.put(_cache.es, ck, {"scores": list(np.asarray(es, dtype=float))})
    return np.asarray(es, dtype=float)


def _install_cached_pipeline() -> None:
    eu.fetch_repo_text = _cached_fetch_repo_text
    eu.zero_shot_scores = _cached_zero_shot
    eu.embedding_similarity_scores = _cached_embedding


# ─────────────────────────────── microservice ────────────────────────────────

def _microservice_healthy(url: str) -> bool:
    try:
        r = requests.get(f"{url}/", timeout=5)
        return r.status_code == 200
    except requests.RequestException:
        return False


def boot_microservice(url: str, log_file: Path, timeout_s: int = 900) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    proc = subprocess.Popen(
        [sys.executable, "app.py"],
        cwd=str(ROOT_DIR / "models"),
        stdout=open(log_file, "w", encoding="utf-8"),
        stderr=subprocess.STDOUT,
    )
    atexit.register(proc.kill)
    start = time.time()
    while time.time() - start < timeout_s:
        if _microservice_healthy(url):
            log(f"[microservice] up on {url} after {time.time() - start:.0f}s")
            return
        if proc.poll() is not None:
            log(f"[microservice] process exited early (code {proc.returncode}); log: {log_file}")
            raise RuntimeError("GE-Lab microservice failed to start — see log above")
        time.sleep(2)
    raise RuntimeError(f"GE-Lab microservice did not become healthy in {timeout_s}s")


# ─────────────────────────────── metrics ─────────────────────────────────────

def _eps() -> float:
    return 1e-9


def per_sdg_table(gt: np.ndarray, pred: np.ndarray):
    gt = np.asarray(gt, float)
    pred = np.asarray(pred, float)
    tp = (gt * pred).sum(axis=0)
    fp = ((1 - gt) * pred).sum(axis=0)
    fn = (gt * (1 - pred)).sum(axis=0)
    tn = ((1 - gt) * (1 - pred)).sum(axis=0)
    e = _eps()
    prec = tp / (tp + fp + e)
    rec = tp / (tp + fn + e)
    f1 = 2 * prec * rec / (prec + rec + e)
    return {
        "n_pos": [int(x) for x in gt.sum(axis=0)],
        "tp": [int(x) for x in tp],
        "fp": [int(x) for x in fp],
        "fn": [int(x) for x in fn],
        "tn": [int(x) for x in tn],
        "precision": [float(x) for x in prec],
        "recall": [float(x) for x in rec],
        "f1": [float(x) for x in f1],
    }


def micro_macro(gt: np.ndarray, pred: np.ndarray) -> dict:
    t = per_sdg_table(gt, pred)
    e = _eps()
    micro_prec = sum(t["tp"]) / (sum(t["tp"]) + sum(t["fp"]) + e)
    micro_rec = sum(t["tp"]) / (sum(t["tp"]) + sum(t["fn"]) + e)
    micro_f1 = 2 * micro_prec * micro_rec / (micro_prec + micro_rec + e)
    return {
        "micro_precision": micro_prec,
        "micro_recall": micro_rec,
        "micro_f1": micro_f1,
        "macro_precision": float(np.mean(t["precision"])),
        "macro_recall": float(np.mean(t["recall"])),
        "macro_f1": float(np.mean(t["f1"])),
    }


def roc_aucs(scores: np.ndarray, gt: np.ndarray) -> list[float]:
    from sklearn.metrics import roc_auc_score
    aucs = []
    for i in range(N_SDGS):
        y = gt[:, i]
        if len(np.unique(y)) < 2:
            aucs.append(None)
            continue
        try:
            aucs.append(float(roc_auc_score(y, scores[:, i])))
        except ValueError:
            aucs.append(None)
    return aucs


def calibration(scores: np.ndarray, gt: np.ndarray, bins: int = 10) -> dict:
    flat_s = scores.ravel()
    flat_g = gt.ravel()
    edges = np.linspace(0.0, 1.0, bins + 1)
    rows = []
    total_abs = 0.0
    total_inst = 0
    for lo, hi in zip(edges[:-1], edges[1:]):
        mask = (flat_s >= lo) & (flat_s < hi)
        n = int(mask.sum())
        if n == 0:
            continue
        emp = float(flat_g[mask].mean())
        center = (lo + hi) / 2
        rows.append({"bin": f"{lo:.2f}-{hi:.2f}", "n": n, "predicted": center,
                     "empirical": round(emp, 4), "gap": round(abs(center - emp), 4)})
        total_abs += n * abs(center - emp)
        total_inst += n
    return {"ece": round(total_abs / max(total_inst, 1), 4), "rows": rows}


def threshold_sweep(scores: np.ndarray, gt: np.ndarray, thresholds: np.ndarray) -> list[dict]:
    rows = []
    for t in thresholds:
        pred = (scores >= t).astype(int)
        mm = micro_macro(gt, pred)
        rows.append({
            "threshold": round(float(t), 4),
            "micro_f1": round(mm["micro_f1"], 4),
            "micro_precision": round(mm["micro_precision"], 4),
            "micro_recall": round(mm["micro_recall"], 4),
            "macro_f1": round(mm["macro_f1"], 4),
            "avg_preds_per_project": round(float(pred.sum(axis=1).mean()), 3),
            "total_preds": int(pred.sum()),
        })
    return rows


def per_sdg_optimal_thresholds(scores: np.ndarray, gt: np.ndarray, grid: np.ndarray) -> list[dict]:
    out = []
    for i in range(N_SDGS):
        best_t, best_f1 = grid[0], -1.0
        for t in grid:
            pred = (scores[:, i] >= t).astype(int)
            tp = (gt[:, i] * pred).sum()
            fp = ((1 - gt[:, i]) * pred).sum()
            fn = (gt[:, i] * (1 - pred)).sum()
            e = _eps()
            p = tp / (tp + fp + e)
            r = tp / (tp + fn + e)
            f1 = 2 * p * r / (p + r + e)
            if f1 > best_f1:
                best_f1, best_t = f1, t
        out.append({"sdg": _SDG_NAMES[i], "n_pos": int(gt[:, i].sum()),
                    "optimal_threshold": round(float(best_t), 3),
                    "f1_at_optimal": round(float(best_f1), 3)})
    return out


# ─────────────────────────────── LLM checks ──────────────────────────────────

def _content_words(text: str) -> set[str]:
    return {w for w in re.findall(r"[A-Za-z][A-Za-z'-]{2,}", text.lower()) if w not in _STOPWORDS}


def _numbers(text: str) -> set[str]:
    return {n for n in _NUMBER_RE.findall(text)}


def summary_quality_check(summary: str, readme: str, desc: str, name: str, topics: list[str]) -> dict:
    source_text = " ".join([readme, desc, name, " ".join(topics)])
    source_words = _content_words(source_text)
    sum_words = _content_words(summary)
    overlap = len(sum_words & source_words) / max(len(sum_words), 1)
    oov = sum_words - source_words
    nums_src = _numbers(source_text)
    nums_sum = _numbers(summary)
    nums_oov = nums_sum - nums_src
    banned_tech = sorted(w for w in _TECH_NOISE if w in summary.lower())
    sdg_leak = bool(_SDG_LEAK_RE.search(summary))
    word_count = len(summary.split())
    sent_count = max(len([s for s in re.split(r"[.!?]+", summary) if s.strip()]), 1)
    relevance = None
    if readme:
        try:
            emb = eu.get_embedder()
            v1 = emb.encode([summary], normalize_embeddings=True)[0]
            v2 = emb.encode([readme[:6000]], normalize_embeddings=True)[0]
            relevance = float(np.dot(v1, v2))
        except Exception:
            pass
    return {
        "word_count": word_count,
        "sentence_count": sent_count,
        "word_overlap_with_source": round(overlap, 3),
        "out_of_vocab_words": sorted(oov)[:25],
        "out_of_vocab_count": len(oov),
        "num_claims_in_summary": sorted(nums_sum),
        "num_claims_in_source": sorted(nums_src),
        "num_claims_not_in_source": sorted(nums_oov),
        "relevance_cosine": relevance,
        "tech_noise_hits": banned_tech,
        "sdg_number_leak": sdg_leak,
        "no_signal_marker": summary.strip().upper() == _NO_SIGNAL or summary.strip().startswith(_NO_SIGNAL),
        "insufficient_doc_marker": summary.strip().lower().startswith("insufficient documentation"),
    }


def summarize_llm_stats(checks: list[dict]) -> dict:
    if not checks:
        return {}
    mean = lambda key: round(float(np.mean([c[key] for c in checks if c[key] is not None])), 4) \
        if any(c[key] is not None for c in checks) else None
    return {
        "n": len(checks),
        "avg_word_count": round(float(np.mean([c["word_count"] for c in checks])), 2),
        "avg_sentence_count": round(float(np.mean([c["sentence_count"] for c in checks])), 2),
        "avg_word_overlap": mean("word_overlap_with_source"),
        "avg_relevance_cosine": mean("relevance_cosine"),
        "avg_oov_count": round(float(np.mean([c["out_of_vocab_count"] for c in checks])), 2),
        "projects_with_oov_gt_40pct": sum(1 for c in checks if c["word_overlap_with_source"] < 0.6),
        "projects_with_sdg_leak": sum(1 for c in checks if c["sdg_number_leak"]),
        "projects_with_tech_noise": sum(1 for c in checks if c["tech_noise_hits"]),
        "projects_with_no_signal": sum(1 for c in checks if c["no_signal_marker"]),
        "projects_with_insufficient_doc": sum(1 for c in checks if c["insufficient_doc_marker"]),
        "projects_with_num_claims": sum(1 for c in checks if c["num_claims_not_in_source"]),
    }


# ─────────────────────────────── pipeline run ────────────────────────────────

def run_project(proj: dict, offline: bool) -> dict:
    url, desc = proj["url"], proj["desc"]
    try:
        result = eu.classify_repo(
            url, threshold=0.0, top_k=N_SDGS, use_ensemble=True, proj_desc=desc
        )
    except ProviderError as e:
        return {"ok": False, "reason": f"ProviderError: {e}"}
    except ValueError as e:
        return {"ok": False, "reason": f"ValueError: {e}"}
    except requests.RequestException as e:
        return {"ok": False, "reason": f"RequestException: {e}"}
    except Exception as e:
        return {"ok": False, "reason": f"{type(e).__name__}: {e}"}

    summary = _summary_by_url.get(url, "")
    raw = _get_raw_sources(url)
    if not summary:
        return {"ok": False, "reason": "empty extracted text (no README/description)"}

    zs, _ = _cached_zero_shot(summary, _SDG_NAMES)
    es = _cached_embedding(summary, sdg_constants.SDG_DESCS)
    scores_all = DEV_ALPHA * zs + (1 - DEV_ALPHA) * es
    top_all = [float(s) for s in scores_all]

    pipeline_score_vec = dict(result["top_all"])
    recomputed = dict(zip(_SDG_NAMES, top_all))
    mismatch = sum(1 for n in _SDG_NAMES
                   if abs(pipeline_score_vec.get(n, -1) - recomputed.get(n, -2)) > 1e-6)

    ranked = sorted(recomputed.items(), key=lambda kv: kv[1], reverse=True)
    preds_prod = [(n, s) for n, s in ranked if s >= DEV_THRESHOLD_MAIN][:DEV_TOP_K]
    preds_50 = [(n, s) for n, s in ranked if s >= DEV_THRESHOLD_CLASSIFY]
    preds_40 = [(n, s) for n, s in ranked if s >= DEV_THRESHOLD_APP]

    return {
        "ok": True,
        "reason": "",
        "summary": summary,
        "summary_len": len(summary.split()),
        "readme_len": len(raw.get("readme", "")),
        "desc_len": len(desc),
        "meta": result["meta"],
        "zs": np.asarray(zs, dtype=float),
        "es": np.asarray(es, dtype=float),
        "scores": scores_all,
        "ranked": ranked,
        "pipeline_mismatch": mismatch,
        "preds_prod": preds_prod,
        "preds_50": preds_50,
        "preds_40": preds_40,
    }


def run_set(projects: list[dict], label: str, offline: bool, delay: float = 10.0) -> tuple[list[dict], np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    all_results = []
    scores_mat = np.zeros((len(projects), N_SDGS))
    zs_mat = np.zeros((len(projects), N_SDGS))
    es_mat = np.zeros((len(projects), N_SDGS))
    gt_mat = np.zeros((len(projects), N_SDGS), dtype=int)
    ok_count = 0
    start = time.time()
    for idx, proj in enumerate(projects):
        global _live_call_made
        _live_call_made = False
        t0 = time.time()
        res = run_project(proj, offline)
        res["_project"] = proj
        all_results.append(res)
        if res["ok"]:
            scores_mat[ok_count] = res["scores"]
            zs_mat[ok_count] = res["zs"]
            es_mat[ok_count] = res["es"]
            if proj["gt"] is not None:
                gt_mat[ok_count] = proj["gt"]
            ok_count += 1
        elapsed = time.time() - t0
        done = idx + 1
        eta = (time.time() - start) / done * (len(projects) - done)
        status = "ok" if res["ok"] else f"FAIL({res['reason'][:40]})"
        log(f"[{label}] {done}/{len(projects)} {proj['name'][:40]:<40} "
            f"{elapsed:5.1f}s {status} eta {eta/60:6.1f}m")
        if delay > 0 and _live_call_made:
            log(f"[{label}] live call made — sleeping {delay:.0f}s to respect rate limits")
            time.sleep(delay)
    return all_results, scores_mat[:ok_count], zs_mat[:ok_count], es_mat[:ok_count], gt_mat[:ok_count]


# ─────────────────────────────── reporting ───────────────────────────────────

def fmt_pct(x) -> str:
    return "n/a" if x is None else f"{x:.3f}"


def print_per_sdg(title: str, t: dict, aucs: list = None) -> None:
    log(f"\n{title}")
    log(f"{'SDG':<42}{'n+':>4}{'TP':>4}{'FP':>4}{'FN':>4}{'Prec':>8}{'Rec':>8}{'F1':>8}{'AUC':>8}")
    for i in range(N_SDGS):
        auc = fmt_pct(aucs[i]) if aucs else "-"
        log(f"{_SDG_NAMES[i][:40]:<42}{t['n_pos'][i]:>4}{t['tp'][i]:>4}{t['fp'][i]:>4}"
            f"{t['fn'][i]:>4}{t['precision'][i]:>8.3f}{t['recall'][i]:>8.3f}{t['f1'][i]:>8.3f}{auc:>8}")


def per_sdg_gate_vector() -> np.ndarray:
    """Per-SDG threshold vector aligned to _SDG_NAMES order, from
    sdg_constants.PER_SDG_THRESHOLDS (alpha-0.7 derivation, see
    docs/EVAL_DPGA_150_RESULTS_alpha07_ANALYSIS.md §6)."""
    return np.array([sdg_constants.PER_SDG_THRESHOLDS[str(i + 1)] for i in range(N_SDGS)], dtype=float)


def print_per_sdg_psdg(title: str, t: dict, gates: np.ndarray, t_ref: dict = None) -> None:
    """Per-SDG metrics at the per-SDG threshold operating point, with the dev
    threshold (0.7) F1 alongside for comparison."""
    log(f"\n{title}")
    log(f"{'SDG':<42}{'thr':>6}{'TP':>4}{'FP':>4}{'FN':>4}{'Prec':>8}{'Rec':>8}{'F1':>8}{'F1@0.7':>8}")
    for i in range(N_SDGS):
        f1_ref = f"{t_ref['f1'][i]:.3f}" if t_ref else "-"
        log(f"{_SDG_NAMES[i][:40]:<42}{gates[i]:>6.2f}{t['tp'][i]:>4}{t['fp'][i]:>4}{t['fn'][i]:>4}"
            f"{t['precision'][i]:>8.3f}{t['recall'][i]:>8.3f}{t['f1'][i]:>8.3f}{f1_ref:>8}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Evaluate UNSDG classifier on real DPG ground truth.")
    ap.add_argument("--dpgs", type=int, default=141, help="max labeled DPGs to evaluate (0 = skip)")
    ap.add_argument("--opensustain", type=int, default=150, help="aux OpenSustain sample size (0 = skip)")
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--cache-dir", type=str, default=str(CACHE_DEFAULT))
    ap.add_argument("--out", type=str, default="", help="report JSON path (default: cache dir)")
    ap.add_argument("--boot-models", action="store_true", help="start the GE-Lab microservice on :9010")
    ap.add_argument("--microservice-url", type=str, default="http://localhost:9010")
    ap.add_argument("--offline", action="store_true", help="use disk cache only; fail on any miss")
    ap.add_argument("--llm-sample", type=int, default=0, help="max projects for LLM-quality checks (0 = all)")
    ap.add_argument("--delay", type=float, default=10.0,
                    help="seconds to sleep after each project that hit a live service (0 = none)")
    ap.add_argument("--smoke", action="store_true", help="tiny quick run (same as small --dpgs/--opensustain)")
    ap.add_argument("-q", "--quiet", action="store_true", help="suppress per-project progress")
    args = ap.parse_args()

    global _cache, log
    if args.quiet:
        def log(msg):  # noqa: F811
            pass

    _cache = DiskCache(args.cache_dir)
    _install_cached_pipeline()

    dpgs_path = ROOT_DIR / "dpgs.csv.xlsx"
    os_path = ROOT_DIR / "OpenSustain.tech-Projects (1).csv.xlsx"
    if not dpgs_path.exists():
        log(f"missing {dpgs_path}")
        return
    if not os_path.exists():
        log(f"missing {os_path}")
        return

    dpgs_all = load_dpgs(dpgs_path)
    exclude_keys = {_norm_key(p["url"]) for p in dpgs_all}
    aux_all = load_opensustain(os_path, args.opensustain or 0, args.seed, exclude_keys)
    dpgs = dpgs_all[:args.dpgs] if args.dpgs else []
    aux = aux_all[:args.opensustain] if args.opensustain else []

    log(f"[data] dpgs.csv.xlsx: {len(dpgs_all)} labeled DPGs ({len(dpgs)} selected)")
    log(f"[data] OpenSustain: {len(aux)} aux projects sampled from {len(aux_all)} supported candidates")
    log(f"[data] labeled positives per SDG: {[int(x) for x in np.array([p['gt'] for p in dpgs_all]).sum(axis=0)]}")

    if not args.offline:
        if not _microservice_healthy(args.microservice_url):
            if args.boot_models:
                log("[microservice] not healthy — booting...")
                boot_microservice(args.microservice_url, Path(args.cache_dir) / "models_boot.log")
            else:
                log("[microservice] NOT running on :9010. Start it (cd models && python app.py) "
                    "or pass --boot-models. Rerun with --offline to reuse cache without it.")
                return

    report: dict = {
        "meta": {
            "script": "eval_dpga_150.py",
            "ts": now(),
            "seed": args.seed,
            "dpgs_selected": len(dpgs),
            "opensustain_selected": len(aux),
            "offline": args.offline,
            "dev_config": {
                "main_threshold": DEV_THRESHOLD_MAIN,
                "classify_threshold": DEV_THRESHOLD_CLASSIFY,
                "app_filter": DEV_THRESHOLD_APP,
                "alpha": DEV_ALPHA,
                "cosine_low": DEV_COSINE_LOW,
                "cosine_high": DEV_COSINE_HIGH,
                "top_k": DEV_TOP_K,
            },
        },
        "dataset": {},
        "production": {},
        "threshold_sweep": {},
        "alpha_sweep": {},
        "reliability": {},
        "pipeline_integrity": {},
        "llm_summary": {},
        "flags": [],
    }

    # ── labeled eval (dpgs.csv.xlsx) ──────────────────────────────────────
    if dpgs:
        log("\n=== RUNNING LABELED EVAL (dpgs.csv.xlsx) ===")
        all_res, scores, zs, es, gt = run_set(dpgs, "dpgs", args.offline, delay=args.delay)
        res = [r for r in all_res if r["ok"]]
        n = len(res)
        report["dataset"]["labeled_ok"] = n
        report["dataset"]["labeled_failed"] = len(dpgs) - n
        report["dataset"]["labeled_failures"] = [
            {"name": r["_project"]["name"], "url": r["_project"]["url"], "reason": r["reason"]}
            for r in all_res if not r["ok"]
        ]

        t = per_sdg_table(gt, (scores >= DEV_THRESHOLD_MAIN).astype(int))
        mm = micro_macro(gt, (scores >= DEV_THRESHOLD_MAIN).astype(int))
        aucs = roc_aucs(scores, gt)
        print_per_sdg(f"PER-SDG METRICS @ dev threshold {DEV_THRESHOLD_MAIN} (top_k {DEV_TOP_K})", t, aucs)
        log(f"\nMICRO  P={mm['micro_precision']:.3f} R={mm['micro_recall']:.3f} F1={mm['micro_f1']:.3f}")
        log(f"MACRO  P={mm['macro_precision']:.3f} R={mm['macro_recall']:.3f} F1={mm['macro_f1']:.3f}")
        report["production"]["threshold_main"] = {
            "per_sdg": t, "micro": {k: round(v, 4) for k, v in mm.items()}, "aucs": aucs,
        }

        for name, thr in (("classify_0.55", DEV_THRESHOLD_CLASSIFY), ("app_0.4", DEV_THRESHOLD_APP)):
            mm_i = micro_macro(gt, (scores >= thr).astype(int))
            log(f"@threshold {thr}: micro F1={mm_i['micro_f1']:.3f} "
                f"(P={mm_i['micro_precision']:.3f} R={mm_i['micro_recall']:.3f}) "
                f"macro F1={mm_i['macro_f1']:.3f}")
            report["production"][name] = {k: round(v, 4) for k, v in mm_i.items()}

        # ── NEW pipeline operating point: per-SDG thresholds ────────────────
        # This is what classify_repo/main() now apply (embedding_url.py's
        # passes_threshold with sdg_constants.PER_SDG_THRESHOLDS). The scores
        # matrix reproduces classify_repo's ensemble scores bit-for-bit, so
        # (scores >= gate) is exactly the new pipeline's selection.
        gates = per_sdg_gate_vector()
        pred_psdg = (scores >= gates).astype(int)
        t_psdg = per_sdg_table(gt, pred_psdg)
        mm_psdg = micro_macro(gt, pred_psdg)
        print_per_sdg_psdg(
            f"PER-SDG METRICS @ per-SDG thresholds (alpha={DEV_ALPHA}; "
            "docs/EVAL_DPGA_150_RESULTS_alpha07_ANALYSIS.md §6)",
            t_psdg, gates, t_ref=t,
        )
        log(f"\nPER-SDG-THRESHOLD MICRO  P={mm_psdg['micro_precision']:.3f} "
            f"R={mm_psdg['micro_recall']:.3f} F1={mm_psdg['micro_f1']:.3f}")
        log(f"PER-SDG-THRESHOLD MACRO  P={mm_psdg['macro_precision']:.3f} "
            f"R={mm_psdg['macro_recall']:.3f} F1={mm_psdg['macro_f1']:.3f}")
        log(f"avg predictions/project @ per-SDG thresholds: {pred_psdg.sum(axis=1).mean():.2f} "
            f"(dev 0.7: {(scores >= DEV_THRESHOLD_MAIN).sum(axis=1).mean():.2f}, "
            f"ground-truth: {gt.sum(axis=1).mean():.2f})")
        report["production"]["per_sdg_thresholds"] = {
            "gates": {f"SDG {i + 1}": round(float(gates[i]), 2) for i in range(N_SDGS)},
            "per_sdg": t_psdg,
            "micro": {k: round(v, 4) for k, v in mm_psdg.items()},
            "avg_preds_per_project": round(float(pred_psdg.sum(axis=1).mean()), 3),
            "avg_gt_per_project": round(float(gt.sum(axis=1).mean()), 3),
        }

        grid = sorted(set(np.round(np.arange(0.0, 1.001, 0.05), 2)) | {0.4, 0.5, 0.7})
        sweep = threshold_sweep(scores, gt, np.array(grid))
        best_micro = max(sweep, key=lambda r: r["micro_f1"])
        best_macro = max(sweep, key=lambda r: r["macro_f1"])
        log(f"\nTHRESHOLD SWEEP (ensemble alpha={DEV_ALPHA})")
        log(f"{'thr':>5}{'microP':>8}{'microR':>8}{'microF1':>8}{'macroF1':>8}{'avgPreds':>9}")
        for r in sweep:
            mark = ""
            if r["threshold"] in (DEV_THRESHOLD_MAIN, DEV_THRESHOLD_CLASSIFY, DEV_THRESHOLD_APP):
                mark = "  <- dev"
            if r["threshold"] == best_micro["threshold"]:
                mark += "  [best micro-F1]"
            log(f"{r['threshold']:>5.2f}{r['micro_precision']:>8.3f}{r['micro_recall']:>8.3f}"
                f"{r['micro_f1']:>8.3f}{r['macro_f1']:>8.3f}{r['avg_preds_per_project']:>9.2f}{mark}")
        log(f"\nOPTIMAL threshold by micro-F1: {best_micro['threshold']:.2f} (F1={best_micro['micro_f1']:.3f})")
        log(f"OPTIMAL threshold by macro-F1: {best_macro['threshold']:.2f} (F1={best_macro['macro_f1']:.3f})")
        report["threshold_sweep"]["grid"] = grid
        report["threshold_sweep"]["rows"] = sweep
        report["threshold_sweep"]["optimal_micro"] = best_micro
        report["threshold_sweep"]["optimal_macro"] = best_macro
        report["threshold_sweep"]["dev_thresholds"] = {
            "0.4": next(r for r in sweep if r["threshold"] == DEV_THRESHOLD_APP),
            "0.5": next(r for r in sweep if r["threshold"] == DEV_THRESHOLD_CLASSIFY),
            "0.7": next(r for r in sweep if r["threshold"] == DEV_THRESHOLD_MAIN),
        }

        fine = np.round(np.arange(0.0, 1.001, 0.01), 2)
        opt_sdg = per_sdg_optimal_thresholds(scores, gt, fine)
        log("\nPER-SDG OPTIMAL THRESHOLDS (1% grid)")
        log(f"{'SDG':<42}{'n+':>4}{'best_t':>8}{'F1':>8}")
        for o in opt_sdg:
            log(f"{o['sdg'][:40]:<42}{o['n_pos']:>4}{o['optimal_threshold']:>8.2f}{o['f1_at_optimal']:>8.3f}")
        report["threshold_sweep"]["per_sdg_optimal"] = opt_sdg

        alphas = [round(a, 1) for a in np.arange(0.0, 1.05, 0.1)]
        alpha_rows = []
        for a in alphas:
            sc = a * zs + (1 - a) * es
            pred = (sc >= DEV_THRESHOLD_MAIN).astype(int)
            mm_a = micro_macro(gt, pred)
            best_t = max(threshold_sweep(sc, gt, np.array(grid)), key=lambda r: r["micro_f1"])
            alpha_rows.append({
                "alpha": a, "micro_f1@0.7": round(mm_a["micro_f1"], 4),
                "macro_f1@0.7": round(mm_a["macro_f1"], 4),
                "best_micro_f1": best_t["micro_f1"], "best_t": best_t["threshold"],
            })
        log("\nALPHA SWEEP (threshold fixed at 0.7)  [alpha=1 -> pure zero-shot, alpha=0 -> pure embedding]")
        log(f"{'alpha':>6}{'microF1@0.7':>12}{'macroF1@0.7':>12}{'bestMicroF1':>12}{'bestT':>7}")
        for r in alpha_rows:
            mark = "  <- dev" if abs(r["alpha"] - DEV_ALPHA) < 1e-9 else ""
            log(f"{r['alpha']:>6.1f}{r['micro_f1@0.7']:>12.3f}{r['macro_f1@0.7']:>12.3f}"
                f"{r['best_micro_f1']:>12.3f}{r['best_t']:>7.2f}{mark}")
        report["alpha_sweep"]["dev_alpha"] = DEV_ALPHA
        report["alpha_sweep"]["rows"] = alpha_rows

        cal = calibration(scores, gt)
        log(f"\nCALIBRATION (pooled SDG-instances) ECE={cal['ece']}")
        for r in cal["rows"]:
            log(f"  score {r['bin']:<12} n={r['n']:<6} predicted={r['predicted']:.2f} "
                f"empirical={r['empirical']:.3f} gap={r['gap']:.3f}")
        report["reliability"]["calibration"] = cal
        report["reliability"]["aucs"] = aucs

        mean_conf_pos = float(scores[gt.astype(bool)].mean()) if gt.any() else None
        mean_conf_neg = float(scores[~gt.astype(bool)].mean()) if (~gt.astype(bool)).any() else None
        report["reliability"]["mean_score_positives"] = mean_conf_pos
        report["reliability"]["mean_score_negatives"] = mean_conf_neg
        log(f"\nRELIABILITY: mean score on true positives={mean_conf_pos:.3f}, "
            f"on negatives={mean_conf_neg:.3f}, avg predictions/project="
            f"{(scores >= DEV_THRESHOLD_MAIN).sum(axis=1).mean():.2f}")

        integrity = {
            "checked": n,
            "pipeline_mismatch_count": sum(1 for r in res if r["pipeline_mismatch"]),
        }
        log(f"\nPIPELINE INTEGRITY: classify_repo scores == recomputed-from-summary scores "
            f"in {integrity['checked'] - integrity['pipeline_mismatch_count']}/{integrity['checked']} projects")
        report["pipeline_integrity"]["labeled"] = integrity

        llm_sample = res[:args.llm_sample] if args.llm_sample else res
        checks = []
        for r in llm_sample:
            raw = _get_raw_sources(r["_project"]["url"])
            c = summary_quality_check(
                r["summary"], raw.get("readme", ""), r["_project"]["desc"],
                raw.get("name", ""), raw.get("topics", []) or [],
            )
            c["project"] = r["_project"]["name"]
            c["url"] = r["_project"]["url"]
            c["summary"] = r["summary"]
            checks.append(c)
            # Flag everything worth a manual audit (see EVAL_DPGA_150_RESULTS.md
            # §8 rec #6): low word-overlap, SDG leaks, and numeric claims that
            # are not present in the source README (hallucination signature).
            # Each flag carries the full summary so the audit needs no network.
            if c["word_overlap_with_source"] < 0.6:
                report["flags"].append({
                    "project": c["project"], "url": c["url"],
                    "issue": "low word-overlap (possible hallucination)",
                    "detail": f"overlap={c['word_overlap_with_source']}, oov={c['out_of_vocab_words'][:6]}",
                    "summary": c["summary"],
                })
            if c["sdg_number_leak"]:
                report["flags"].append({
                    "project": c["project"], "url": c["url"],
                    "issue": "SDG number leaked into summary",
                    "detail": "summary names an SDG/Goal number verbatim",
                    "summary": c["summary"],
                })
            if c["num_claims_not_in_source"]:
                report["flags"].append({
                    "project": c["project"], "url": c["url"],
                    "issue": "numeric claims not in source (possible hallucination)",
                    "detail": ("claims_in_summary=" + str(c["num_claims_in_summary"]) +
                               ", not_in_source=" + str(c["num_claims_not_in_source"])),
                    "summary": c["summary"],
                })
        stats = summarize_llm_stats(checks)
        log(f"\nLLM SUMMARY QUALITY (n={stats.get('n', 0)})")
        for k in ("avg_word_count", "avg_sentence_count", "avg_word_overlap",
                  "avg_relevance_cosine", "avg_oov_count", "projects_with_sdg_leak",
                  "projects_with_tech_noise", "projects_with_num_claims",
                  "projects_with_no_signal", "projects_with_insufficient_doc"):
            log(f"  {k:<32}{stats.get(k)}")
        report["llm_summary"]["labeled"] = {"stats": stats, "per_project": checks}

        preds_csv = []
        for r in res:
            proj = r["_project"]
            preds_csv.append({
                "name": proj["name"], "url": proj["url"],
                "ground_truth": ",".join(str(i + 1) for i, v in enumerate(proj["gt"]) if v),
                "preds_at_0.7": ",".join(n.split(":")[0] for n, _ in r["preds_prod"]),
                "preds_per_sdg": ",".join(
                    str(i + 1) for i in range(N_SDGS) if r["scores"][i] >= gates[i]
                ),
                "top1": r["ranked"][0][0],
                "summary_len": r["summary_len"],
                **{f"score_sdg{i+1}": round(float(r["scores"][i]), 4) for i in range(N_SDGS)},
            })
        report["predictions"] = preds_csv

    # ── aux eval (OpenSustain sample, no ground truth) ────────────────────
    if aux:
        log("\n=== RUNNING AUX EVAL (OpenSustain.tech sample) ===")
        all_res, scores, zs, es, _ = run_set(aux, "aux", args.offline, delay=args.delay)
        res = [r for r in all_res if r["ok"]]
        n = len(res)
        report["dataset"]["aux_ok"] = n
        report["dataset"]["aux_failed"] = len(aux) - n
        report["dataset"]["aux_failures"] = [
            {"name": r["_project"]["name"], "url": r["_project"]["url"], "reason": r["reason"]}
            for r in all_res if not r["ok"]
        ]
        zero_preds = int(((scores >= DEV_THRESHOLD_MAIN).sum(axis=1) == 0).sum())
        log(f"\nAUX SCORE DISTRIBUTION (no ground truth) — {n} projects")
        log(f"  max score/project:   mean={scores.max(axis=1).mean():.3f} "
            f"p25={np.percentile(scores.max(axis=1), 25):.3f} "
            f"median={np.percentile(scores.max(axis=1), 50):.3f} "
            f"p90={np.percentile(scores.max(axis=1), 90):.3f}")
        log(f"  predictions/project @0.7: mean={(scores >= DEV_THRESHOLD_MAIN).sum(axis=1).mean():.2f}")
        log(f"  projects with zero preds @0.7: {zero_preds}")
        gates_aux = per_sdg_gate_vector()
        psdg_preds = (scores >= gates_aux).sum(axis=1)
        log(f"  predictions/project @ per-SDG thresholds: mean={psdg_preds.mean():.2f}")
        log(f"  projects with zero preds @ per-SDG thresholds: {(psdg_preds == 0).sum()}")
        report["aux"] = {
            "n": n,
            "max_score_stats": {
                "mean": round(float(scores.max(axis=1).mean()), 4),
                "p25": round(float(np.percentile(scores.max(axis=1), 25)), 4),
                "median": round(float(np.percentile(scores.max(axis=1), 50)), 4),
                "p90": round(float(np.percentile(scores.max(axis=1), 90)), 4),
            },
            "preds_per_project_@0.7_mean": round(float((scores >= DEV_THRESHOLD_MAIN).sum(axis=1).mean()), 3),
            "zero_preds_count": zero_preds,
            "preds_per_project_@per_sdg_mean": round(float(psdg_preds.mean()), 3),
            "zero_preds_count_per_sdg": int((psdg_preds == 0).sum()),
        }

        llm_sample = res[:args.llm_sample] if args.llm_sample else res
        checks = []
        for r in llm_sample:
            raw = _get_raw_sources(r["_project"]["url"])
            c = summary_quality_check(
                r["summary"], raw.get("readme", ""), r["_project"]["desc"],
                raw.get("name", ""), raw.get("topics", []) or [],
            )
            c["project"] = r["_project"]["name"]
            c["url"] = r["_project"]["url"]
            checks.append(c)
        report["llm_summary"]["aux"] = {"stats": summarize_llm_stats(checks), "per_project": checks}

    ts = time.strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.cache_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = Path(args.out) if args.out else out_dir / f"report_{ts}.json"
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    log(f"\n[done] report -> {out_path}")
    if report.get("predictions"):
        csv_path = out_dir / f"predictions_{ts}.csv"
        cols = list(report["predictions"][0].keys())
        csv_path.write_text(
            "\n".join([",".join(cols)] + [
                ",".join(str(p.get(c, "")) for c in cols) for p in report["predictions"]
            ]), encoding="utf-8")
        log(f"[done] predictions -> {csv_path}")
    if report["flags"]:
        log(f"[warning] {len(report['flags'])} LLM-summary flags recorded (see report['flags'])")
        for f in report["flags"][:10]:
            log(f"    - {f['project'][:45]:<45} {f['issue']}")


if __name__ == "__main__":
    main()